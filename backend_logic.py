# backend_logic.py
import os
import re
import logging
import configparser
from datetime import datetime
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.chart import PieChart, BarChart, Reference
from PySide6.QtCore import QObject, Signal

logger = logging.getLogger(__name__)

# Разделенные заголовки для разных вкладок и служебная колонка
HEADERS_MAIN = [
    'Название ПК', 'Имя файла', 'ОС', 'Процессор', 'Сокет',
    'Материнская плата', 'Видеоадаптер', 'Монитор', 'Принтеры', 'Объем ОЗУ',
    'Кол-во плашек ОЗУ', 'Свободно слотов ОЗУ', 'Модели плашек ОЗУ',
    'Дисковые накопители', 'SMART Статус', 'Дата BIOS', '_RAW_DATA'
]
HEADERS_NETWORK = [
    'Название ПК', 'Имя файла', 'Локальный IP', 'MAC-адрес'
]
HEADERS_ANALYSIS = ['Имя файла', 'Название ПК', 'Ключевые проблемы', 'Рекомендация']
CRITICAL_SMART_ATTRIBUTES = {'05', 'C5', 'C6'}

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

class ExcelReaderWorker(QObject):
    log_message = Signal(str, str)
    progress_update = Signal(int, int)
    result_ready = Signal(dict)
    finished = Signal(str)

    def __init__(self, filepath):
        super().__init__()
        self.filepath = filepath

    def run(self): # <- это внутри ExcelReaderWorker
        try:
            self.log_message.emit(f"Загрузка данных из {os.path.basename(self.filepath)}...", "info")
            wb = load_workbook(self.filepath)
            # ВНИМАНИЕ: get_sheet_by_name устарел, используем wb['название']
            ws = wb["Все данные"]
            
            headers = [cell.value for cell in ws[1]]
            
            try:
                headers.index('_RAW_DATA')
            except ValueError:
                self.log_message.emit("Ошибка: В файле Excel отсутствует служебная колонка. Запустите анализ заново.", "error")
                self.finished.emit(""); return
            
            total_rows = ws.max_row - 1
            if total_rows <= 0:
                self.log_message.emit("Файл Excel пуст.", "warning")
                self.finished.emit(self.filepath); return

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                data_row = dict(zip(headers, row))
                raw_data_str = data_row.get('_RAW_DATA')
                if not raw_data_str: continue

                try:
                    parts = raw_data_str.split('|')
                    data_row['category'] = int(parts[0])
                    data_row['_internal_smart_status'] = parts[1]
                    data_row['SMART Проблемы'] = parts[2:]
                except (IndexError, ValueError):
                    continue
                
                self.result_ready.emit(data_row)
                self.progress_update.emit(row_idx, total_rows)

            self.finished.emit(self.filepath)
        except Exception as e:
            self.log_message.emit(f"Ошибка при чтении Excel: {e}", "error")
            logger.error(f"Ошибка при чтении Excel: {e}", exc_info=True)
            self.finished.emit("")


class AidaWorker(QObject):
    log_message = Signal(str, str)
    progress_update = Signal(int, int)
    result_ready = Signal(dict)
    finished = Signal(str)

    def __init__(self, reports_dir, config):
        super().__init__()
        self.reports_dir = reports_dir
        self.config = config
        self.is_running = True

    def run(self):
        try:
            output_file = self.config.get('Settings', 'output_filename', fallback='system_analysis.xlsx')
            report_files = [f for f in os.listdir(self.reports_dir) if f.lower().endswith(('.htm', '.html'))]
            if not report_files:
                self.log_message.emit("В указанной папке не найдено файлов отчетов .htm/.html.", "warning")
                self.finished.emit(""); return

            self.log_message.emit(f"Найдено отчетов: {len(report_files)}", "info")
            all_reports_data = []
            total_files = len(report_files)
            for i, filename in enumerate(report_files):
                if not self.is_running: break
                self.progress_update.emit(i + 1, total_files)
                file_path = os.path.join(self.reports_dir, filename)
                data = self.parse_aida_report(file_path)
                if data:
                    category, problems = self.analyze_system(data)
                    data['category'] = category
                    data['problems'] = problems
                    self.result_ready.emit(data)
                    all_reports_data.append(data)
            
            if not self.is_running:
                self.log_message.emit("Процесс анализа был прерван.", "warning"); self.finished.emit(""); return

            self.log_message.emit("Сортировка данных...", "info")
            all_reports_data.sort(key=lambda item: natural_sort_key(item.get('Имя файла', '')))
            self.log_message.emit("Сохранение в Excel...", "info")
            self.write_to_excel(all_reports_data, output_file)
            self.finished.emit(output_file)

        except Exception as e:
            self.log_message.emit(f"КРИТИЧЕСКАЯ ОШИБКА: {e}", "error")
            logger.error(f"Критическая ошибка в потоке: {e}", exc_info=True)
            self.finished.emit("")

    def find_value_by_label(self, search_area, label_text):
        if not search_area:
            return None
        try:
            candidates = search_area.find_all(
                lambda tag: tag.name == 'td' and label_text in tag.get_text() and not tag.find('td')
            )
            if not candidates:
                return None
            label_td = candidates[-1]
            value_td = label_td.find_next_sibling('td')
            if not value_td:
                value_td = label_td.find_next('td')
            if value_td:
                link = value_td.find('a')
                if link:
                    return link.get_text(strip=True)
                return value_td.get_text(strip=True)
            return None
        except Exception as e:
            logger.error(f"Ошибка в find_value_by_label для '{label_text}': {e}", exc_info=True)
            return None

    def parse_smart_data(self, soup):
        smart_section = soup.find('a', {'name': 'smart'})
        if not smart_section: return "Не найден", []
        all_problems = []; has_bad_status=False; has_ok_status=False
        temp_warn = self.config.getint('SMART', 'temp_warning_celsius', fallback=45); power_on_warn = self.config.getint('SMART', 'power_on_warning_hours', fallback=30000); power_cycle_warn = self.config.getint('SMART', 'power_cycle_warning_count', fallback=10000); read_error_warn = self.config.getint('SMART', 'read_error_warning_rate', fallback=1000000)
        drive_tables = smart_section.find_next_siblings('table')
        for table in drive_tables:
            header = table.find('td', class_='dt')
            if not header: continue
            drive_name = header.get_text(strip=True).strip('[] ').split(' ')[0]
            for row in table.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) > 3:
                    attr_id = cells[2].get_text(strip=True); attr_desc = cells[3].get_text(strip=True); raw_data_str = cells[-2].get_text(strip=True)
                    try:
                        raw_data_val = int(raw_data_str, 16) if 'x' in raw_data_str else int(raw_data_str)
                        if attr_id in CRITICAL_SMART_ATTRIBUTES and raw_data_val > 0: has_bad_status = True; all_problems.append(f"{drive_name}: {attr_desc} ({attr_id}) > 0")
                        elif attr_id == 'C2' and raw_data_val > temp_warn: has_ok_status = True; all_problems.append(f"{drive_name}: Высокая t° ({raw_data_val}°C)")
                        elif attr_id == '09' and raw_data_val > power_on_warn: has_ok_status = True; all_problems.append(f"{drive_name}: Большой налет часов")
                        elif attr_id == '0C' and raw_data_val > power_cycle_warn: has_ok_status = True; all_problems.append(f"{drive_name}: Много вкл/выкл")
                        elif attr_id == '01' and raw_data_val > read_error_warn: has_ok_status = True; all_problems.append(f"{drive_name}: Много ошибок чтения")
                    except (ValueError, IndexError): continue
        if has_bad_status: return "BAD", all_problems
        if has_ok_status: return "OK", all_problems
        return "GOOD", []

    def analyze_system(self, data):
        problems = []
        smart_status = data.get('SMART Статус', "Не найден")
        if smart_status == "BAD":
            return 1, "; ".join(data.get('SMART Проблемы', ["Критическая ошибка диска"]))
        if smart_status == "OK":
            problems.extend(data.get('SMART Проблемы', []))
        
        bios_age_limit = self.config.getint('Analysis', 'bios_age_limit_years', fallback=5)
        ram_critical_gb = self.config.getfloat('Analysis', 'ram_critical_gb', fallback=3.8)
        ram_upgrade_gb = self.config.getfloat('Analysis', 'ram_upgrade_gb', fallback=7.8)
        disk_c_critical_gb = self.config.getfloat('Analysis', 'disk_c_critical_gb', fallback=15.0)

        os_ver, cpu, socket = data.get('ОС'), data.get('Процессор'), data.get('Сокет')
        ram_gb_str = re.search(r'(\d+)\s*МБ', data.get('Объем ОЗУ') or '')
        ram_gb = int(ram_gb_str.group(1)) / 1024 if ram_gb_str else 0
        gpu_driver = data.get('Видеоадаптер')
        bios_date_str = data.get('Дата BIOS')
        has_ssd = any(k in (data.get('Дисковые накопители') or '').lower() for k in ['ssd', 'nvme', 'snv'])
        disk_c_free_gb = data.get('Disk_C_Free_GB', -1)

        is_critical = False
        if socket and 'LGA775' in socket:
            is_critical=True; problems.append("Очень старый сокет (LGA775)")
        if ram_gb > 0 and ram_gb < ram_critical_gb:
            is_critical=True; problems.append(f"Критически мало ОЗУ ({ram_gb:.1f} ГБ)")
        if os_ver and 'Windows 7' in os_ver and cpu and any(p in cpu for p in ['G3','G1','E5','i3-2','i3-3','FX']):
            is_critical=True; problems.append("Устаревшая ОС на очень слабом ЦП")
        
        if is_critical:
            return 1, "; ".join(problems)

        is_upgrade_needed = len(problems) > 0
        if os_ver and 'Windows 7' in os_ver:
            is_upgrade_needed=True; problems.append("Устаревшая ОС Windows 7")
        if ram_critical_gb <= ram_gb < ram_upgrade_gb:
            is_upgrade_needed=True; problems.append(f"Недостаточно ОЗУ ({ram_gb:.1f} ГБ)")
        if not has_ssd:
            is_upgrade_needed=True; problems.append("Отсутствует SSD")
        if gpu_driver and 'Microsoft Basic Display Adapter' in gpu_driver:
            is_upgrade_needed=True; problems.append("Не установлен видеодрайвер")
        if disk_c_free_gb != -1 and disk_c_free_gb < disk_c_critical_gb:
            is_upgrade_needed=True; problems.append(f"Мало места на диске C: ({disk_c_free_gb:.1f} ГБ)")

        if bios_date_str:
            try:
                date_match = re.search(r'(\d{2}/\d{2}/\d{2,4})', bios_date_str)
                if date_match:
                    date_str_extracted = date_match.group(1)
                    bios_date_format = '%m/%d/%Y' if int(date_match.group(1).split('/')[0]) <= 12 and int(date_match.group(1).split('/')[1]) <= 31 else '%d/%m/%Y'
                    if len(date_str_extracted.split('/')[-1]) == 2:
                        bios_date_format = bios_date_format.replace('Y', 'y')
                    bios_date = datetime.strptime(date_str_extracted, bios_date_format)
                    if (datetime.now() - bios_date).days > 365 * bios_age_limit:
                        is_upgrade_needed=True; problems.append(f"BIOS старше {bios_age_limit} лет")
            except (ValueError, IndexError): pass

        if is_upgrade_needed:
            return 2, "; ".join(problems)

        return 3, "Состояние хорошее"
    
    def parse_aida_report(self, file_path):
        self.log_message.emit(f"Обработка: {os.path.basename(file_path)}", "info")
        try:
            with open(file_path, 'r', encoding='windows-1251', errors='ignore') as f:
                html_content = f.read()
            soup = BeautifulSoup(html_content, 'lxml')
            data = {'Имя файла': os.path.basename(file_path)}

            summary_section = soup.find('a', {'name': 'summary'})
            summary_table = summary_section.find_next('table') if summary_section else None

            if summary_table:
                data['Название ПК'] = self.find_value_by_label(summary_table, 'Имя компьютера') or ''
                data['ОС'] = self.find_value_by_label(summary_table, 'Операционная система') or ''
                data['Процессор'] = self.find_value_by_label(summary_table, 'Тип ЦП') or ''
                data['Материнская плата'] = self.find_value_by_label(summary_table, 'Системная плата') or ''
                data['Видеоадаптер'] = self.find_value_by_label(summary_table, 'Видеоадаптер') or ''
                data['Монитор'] = self.find_value_by_label(summary_table, 'Монитор') or ''
                data['Объем ОЗУ'] = self.find_value_by_label(summary_table, 'Системная память') or ''
                data['Локальный IP'] = self.find_value_by_label(summary_table, 'Первичный адрес IP') or ''
                data['MAC-адрес'] = self.find_value_by_label(summary_table, 'Первичный адрес MAC') or ''
                disk_c_text = self.find_value_by_label(summary_table, 'C: (NTFS)')
                if disk_c_text:
                    match = re.search(r'\((\d+)\s*МБ\s*свободно\)', disk_c_text)
                    if match:
                        free_mb = int(match.group(1))
                        data['Disk_C_Free_GB'] = round(free_mb / 1024, 1)
                ram_candidates = summary_table.find_all(lambda tag: tag.name == 'td' and re.search(r'DIMM\d', tag.get_text()) and not tag.find('td'))
                ram_models = [re.sub(r'\s+\(.*\)', '', label.find_next('td').get_text(strip=True)) for label in ram_candidates if label.find_next('td')]
                data['Модели плашек ОЗУ'] = "; ".join(filter(None, ram_models)) or 'Не найдено'
                disk_candidates = summary_table.find_all(lambda tag: tag.name == 'td' and 'Дисковый накопитель' in tag.get_text() and not tag.find('td'))
                disk_models = [label.find_next('td').get_text(strip=True) for label in disk_candidates if label.find_next('td')]
                data['Дисковые накопители'] = "; ".join(disk_models) or 'Не найдено'
                printer_candidates = summary_table.find_all(lambda tag: tag.name == 'td' and 'Принтер' in tag.get_text() and not tag.find('td'))
                all_printers = [label.find_next('td').get_text(strip=True) for label in printer_candidates if label.find_next('td')]
                system_printers_keys = ['Fax', 'Microsoft Print to PDF', 'XPS', 'OneNote', 'AnyDesk']
                physical_printers = [p for p in all_printers if not any(key in p for key in system_printers_keys)]
                data['Принтеры'] = "; ".join(physical_printers) or 'Не найдено'

            mobo_section = soup.find('a', {'name': 'motherboard'})
            mobo_table = mobo_section.find_next('table') if mobo_section else summary_table
            data['Сокет'] = self.find_value_by_label(mobo_table, 'Разъёмы для ЦП') or ''

            bios_section = soup.find('a', {'name': 'bios'})
            bios_table = bios_section.find_next('table') if bios_section else summary_table
            data['Дата BIOS'] = self.find_value_by_label(bios_table, 'Дата BIOS системы') or ''

            data['SMART Статус'], data['SMART Проблемы'] = self.parse_smart_data(soup)
            
            installed_sticks_count = data['Модели плашек ОЗУ'].count(';') + 1 if data.get('Модели плашек ОЗУ') and data['Модели плашек ОЗУ'] != 'Не найдено' else 0
            total_ram_slots = 0
            if installed_sticks_count > 0:
                html_content_lower = html_content.lower()
                if installed_sticks_count >= 3: 
                    total_ram_slots = 4
                elif 'so-dimm' in html_content_lower:
                    total_ram_slots = 2
                else: 
                    has_high_slots = 'dimm3' in html_content_lower or 'dimm4' in html_content_lower
                    if has_high_slots: total_ram_slots = 4
                    else: total_ram_slots = 2
            data['Кол-во плашек ОЗУ'] = installed_sticks_count
            data['Свободно слотов ОЗУ'] = total_ram_slots - installed_sticks_count if total_ram_slots > 0 else 'Неизвестно'

            all_headers = list(set(HEADERS_MAIN + HEADERS_NETWORK))
            for key in all_headers:
                if key not in data and key != '_RAW_DATA':
                    data[key] = 'Не найдено'
            
            return data
        except Exception as e:
            logger.error(f"КРИТИЧЕСКАЯ ОШИБКА ПАРСИНГА в {os.path.basename(file_path)}: {e}", exc_info=True)
            return None

    def write_to_excel(self, data_list, filename):
        if not data_list: return
        wb = Workbook()
        header_font=Font(bold=True,color="FFFFFF",name='Calibri',size=11); header_fill=PatternFill(start_color="4F81BD",end_color="4F81BD",fill_type="solid"); header_alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); thin_border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')); data_font=Font(name='Calibri',size=11); data_alignment=Alignment(vertical='center',wrap_text=True,horizontal='left'); cat1_fill=PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid"); cat1_font=Font(color="9C0006",name='Calibri',size=11); cat2_fill=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid"); cat2_font=Font(color="9C6500",name='Calibri',size=11); cat3_fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid"); smart_bad_font=Font(bold=True,color="9C0006")
        
        ws_main = wb.active
        ws_main.title = "Все данные"
        
        # <<< ИЗМЕНЕНИЕ: Создаем единый список ВСЕХ заголовков для сохранения >>>
        all_headers_to_save = sorted(list(set(HEADERS_MAIN + HEADERS_NETWORK)))
        # Перемещаем _RAW_DATA в конец, если он там не оказался
        if '_RAW_DATA' in all_headers_to_save:
            all_headers_to_save.remove('_RAW_DATA')
            all_headers_to_save.append('_RAW_DATA')

        ws_main.append(all_headers_to_save)
        
        for cell in ws_main[1]: cell.font=header_font; cell.fill=header_fill; cell.alignment=header_alignment; cell.border=thin_border
        
        for row_idx, data_row in enumerate(data_list, start=2):
            smart_problems_str = "; ".join(data_row.get('SMART Проблемы', [])) if data_row.get('SMART Проблемы') else data_row.get('SMART Статус', 'OK')
            
            data_row_display = data_row.copy()
            data_row_display['SMART Статус'] = smart_problems_str
            raw_data_string = f"{data_row.get('category', 3)}|{data_row.get('SMART Статус', 'GOOD')}|" + "|".join(data_row.get('SMART Проблемы', []))
            data_row_display['_RAW_DATA'] = raw_data_string
            
            row_values = [data_row_display.get(h, '') for h in all_headers_to_save]
            ws_main.append(row_values)
            
            row_fill, row_font = {1:(cat1_fill,cat1_font), 2:(cat2_fill,cat2_font), 3:(cat3_fill,data_font)}.get(data_row.get('category'),(None,data_font))
            for cell in ws_main[row_idx]:
                cell.font = row_font; cell.alignment = data_alignment; cell.border = thin_border
                if row_fill: cell.fill = row_fill
            
            try: # Защита от отсутствия колонки в старом формате
                smart_cell_index = all_headers_to_save.index('SMART Статус') + 1
                smart_cell = ws_main.cell(row=row_idx, column=smart_cell_index)
                if data_row.get('SMART Статус') == 'BAD': smart_cell.font = smart_bad_font
            except ValueError:
                pass
        
        raw_data_col_letter = get_column_letter(all_headers_to_save.index('_RAW_DATA') + 1)
        ws_main.column_dimensions[raw_data_col_letter].hidden = True

        for col_idx in range(1, ws_main.max_column):
            if ws_main.column_dimensions[get_column_letter(col_idx)].hidden: continue
            column = ws_main.column_dimensions[get_column_letter(col_idx)]
            max_length = 0
            for cell in ws_main[get_column_letter(col_idx)]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 3)
            column.width = min(adjusted_width, 60)
        ws_main.column_dimensions['A'].width = 25
        ws_main.freeze_panes = 'A2'
        
        ws_analysis = wb.create_sheet("Рекомендации"); cat_font = Font(bold=True, name='Calibri', size=12)
        for cat_num, cat_name, cat_fill in [(1, 'Категория 1: Полная замена', cat1_fill), (2, 'Категория 2: Частичный апгрейд', cat2_fill)]:
            ws_analysis.append([cat_name]); merged_cell_header=ws_analysis.cell(row=ws_analysis.max_row,column=1); merged_cell_header.font=cat_font; merged_cell_header.fill=cat_fill; ws_analysis.merge_cells(start_row=ws_analysis.max_row, start_column=1, end_row=ws_analysis.max_row, end_column=len(HEADERS_ANALYSIS)); ws_analysis.append(HEADERS_ANALYSIS)
            for cell in ws_analysis[ws_analysis.max_row]: cell.font=header_font; cell.fill=header_fill; cell.alignment=header_alignment; cell.border=thin_border
            for data in data_list:
                if data.get('category') == cat_num:
                    rec = ""
                    if data.get('SMART Статус') == "BAD": rec = "ЗАМЕНА ДИСКА!"
                    elif cat_num == 1: rec = "Полная замена"
                    else:
                        rec_list = []
                        problems_str = data.get('problems', '').lower()
                        if 'ос' in problems_str or 'windows 7' in problems_str: rec_list.append("Обновить ОС")
                        if 'ssd' in problems_str: rec_list.append("Установить SSD")
                        if 'озу' in problems_str: rec_list.append("Добавить ОЗУ")
                        if 'видеодрайвер' in problems_str: rec_list.append("Установить видеодрайвер")
                        if 'bios' in problems_str: rec_list.append("Обновить BIOS (опционально)")
                        if 'мало места на диске' in problems_str: rec_list.append("Очистить диск C:")
                        rec = ", ".join(rec_list) if rec_list else "Частичный апгрейд"
                    ws_analysis.append([data['Имя файла'], data['Название ПК'], data.get('problems', ''), rec])
            ws_analysis.append([])
        for col_idx in range(1, ws_analysis.max_column + 1):
            max_length = max((len(str(cell.value)) for cell in ws_analysis[get_column_letter(col_idx)] if cell.value and not isinstance(cell, MergedCell)), default=0)
            if col_idx in [1, 3, 4]: max_length = max(max_length, 35) 
            ws_analysis.column_dimensions[get_column_letter(col_idx)].width = min((max_length + 2), 100)
            
        ws_dash = wb.create_sheet("Дашборд"); cat_counts = {'Требуют замены (BAD)':0, 'Нужен апгрейд (OK)':0, 'В порядке (GOOD)':0}; ram_dist = {'< 4 ГБ':0, '4-8 ГБ':0, '8-16 ГБ':0, '> 16 ГБ':0}
        for data in data_list:
            category = data.get('category', 3)
            cat_counts[{1:'Требуют замены (BAD)', 2:'Нужен апгрейд (OK)', 3:'В порядке (GOOD)'}[category]] += 1
            ram_text = data.get('Объем ОЗУ', '0 MB'); ram_gb_match = re.search(r'(\d+)\s*МБ', ram_text or '')
            if ram_gb_match:
                ram_gb = int(ram_gb_match.group(1))/1024
                if ram_gb < 4: ram_dist['< 4 ГБ']+=1
                elif ram_gb < 8: ram_dist['4-8 ГБ']+=1
                elif ram_gb < 16: ram_dist['8-16 ГБ']+=1
                else: ram_dist['> 16 ГБ']+=1
        pie_data_start_row=2; ws_dash.append(['Категория', 'Количество']); [ws_dash.append([n,c]) for n,c in cat_counts.items()]; pie_data_end_row = ws_dash.max_row; ws_dash.append([])
        bar_data_start_row=ws_dash.max_row+1; ws_dash.append(['Объем ОЗУ','Количество']); [ws_dash.append([n,c]) for n,c in ram_dist.items()]; bar_data_end_row=ws_dash.max_row
        pie=PieChart(); pie.title="Состояние парка компьютеров"; labels=Reference(ws_dash,min_col=1,min_row=pie_data_start_row,max_row=pie_data_end_row); data=Reference(ws_dash,min_col=2,min_row=pie_data_start_row-1,max_row=pie_data_end_row)
        pie.add_data(data,titles_from_data=True); pie.set_categories(labels); ws_dash.add_chart(pie, "D2")
        bar=BarChart(); bar.title="Распределение по объему ОЗУ"; bar.legend=None; labels=Reference(ws_dash,min_col=1,min_row=bar_data_start_row,max_row=bar_data_end_row); data=Reference(ws_dash,min_col=2,min_row=bar_data_start_row-1,max_row=bar_data_end_row)
        bar.add_data(data,titles_from_data=True); bar.set_categories(labels); ws_dash.add_chart(bar, "D20")
        ws_dash.column_dimensions['A'].width=25; ws_dash.column_dimensions['B'].width=15
        
        try:
            wb.save(filename)
            self.log_message.emit(f"Готово! Данные и анализ сохранены в {filename}", "info")
            logger.info(f"Файл Excel '{filename}' успешно сохранен.")
        except IOError as e:
            self.log_message.emit(f"Ошибка: Не удалось записать в файл {filename}. Ошибка: {e}", "error")
            logger.error(f"Ошибка ввода/вывода при сохранении Excel-файла '{filename}': {e}", exc_info=True)