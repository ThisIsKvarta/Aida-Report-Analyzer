# logic/excel_handler.py
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from collections import Counter
from datetime import datetime
import re

from utils.constants import HEADERS_MAIN, HEADERS_NETWORK, HEADERS_ANALYSIS

logger = logging.getLogger(__name__)

def _calculate_statistics(data_list):
    stats = {'total_pcs': len(data_list), 'cat1_critical': 0, 'cat2_upgrade': 0, 'cat3_ok': 0,
             'problem_counts': Counter(), 'bios_dates': [], 'average_bios_age_years': 'N/A', 'top_5_critical': []}
    if not data_list: return stats
    for data in data_list:
        category = data.get('category', 3)
        if category == 1: stats['cat1_critical'] += 1
        elif category == 2: stats['cat2_upgrade'] += 1
        else: stats['cat3_ok'] += 1
        if problems_str := data.get('problems', ''):
            problem_list = [p.strip() for p in problems_str.split(';') if "состояние хорошее" not in p.lower() and p.strip()]
            stats['problem_counts'].update(problem_list)
        if bios_date_str := data.get('Дата BIOS'):
            if date_match := re.search(r'(\d{2}/\d{2}/\d{2,4})', bios_date_str):
                try:
                    p = re.split(r'[/.-]', date_match.group(1)); y = int(p[2]) if int(p[2]) > 1900 else int(p[2]) + 2000
                    stats['bios_dates'].append(datetime(y, int(p[0]), int(p[1])))
                except: pass
    if stats['bios_dates']:
        total_days = sum([(datetime.now() - d).days for d in stats['bios_dates']])
        stats['average_bios_age_years'] = round(total_days / len(stats['bios_dates']) / 365.25, 1)
    stats['top_5_critical'] = sorted(data_list, key=lambda x: x.get('category', 3))[:5]
    return stats

def write_to_excel(data_list, filename, log_emitter):
    if not data_list: log_emitter("Нет данных для экспорта в Excel.", "warning"); return
    wb = Workbook()
    log_emitter("Расчет статистики для дашборда...", "info"); stats = _calculate_statistics(data_list)
    
    header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11); header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    kpi_title_font = Font(name='Calibri', size=11, bold=True, color="595959"); kpi_alignment = Alignment(horizontal='center', vertical='center')
    section_title_font = Font(name='Calibri', size=14, bold=True)

    ws_dash = wb.active; ws_dash.title = "Дашборд"
    ws_dash['B2'] = "Ключевые показатели 'Здоровья' Компьютерного Парка"; ws_dash['B2'].font = Font(name='Calibri', size=18, bold=True)
    ws_dash.merge_cells('B2:K2'); ws_dash.row_dimensions[2].height = 30

    kpi_data = [("B4", "Всего ПК на учете", stats['total_pcs'], "000000"), ("E4", "Требуют ЗАМЕНЫ", stats['cat1_critical'], "9C0006"),
                ("H4", "Требуют АПГРЕЙДА", stats['cat2_upgrade'], "9C6500")]
    for cell_ref, title, value, color in kpi_data:
        start_cell = ws_dash[cell_ref]; min_row, min_col = start_cell.row, start_cell.column
        title_cell = ws_dash.cell(row=min_row, column=min_col, value=title); ws_dash.merge_cells(start_row=min_row, start_column=min_col, end_row=min_row, end_column=min_col + 1)
        title_cell.font = kpi_title_font; title_cell.alignment = kpi_alignment
        value_cell = ws_dash.cell(row=min_row + 1, column=min_col, value=value); ws_dash.merge_cells(start_row=min_row + 1, start_column=min_col, end_row=min_row + 2, end_column=min_col + 1)
        value_cell.font = Font(name='Calibri', size=24, bold=True, color=color); value_cell.alignment = kpi_alignment
        for r in range(min_row, min_row + 3):
            for c in range(min_col, min_col + 2): ws_dash.cell(row=r, column=c).border = thin_border
    
    ws_dash['B8'] = 'Графический анализ'; ws_dash['B8'].font = section_title_font; ws_dash.merge_cells('B8:K8')

    pie = PieChart(); pie.title = "Состояние парка"; pie.height = 10; pie.width = 13
    chart_data_rows = [['Категория', 'Количество'], ['В порядке', stats['cat3_ok']], ['Нужен апгрейд', stats['cat2_upgrade']], ['Критическое состояние', stats['cat1_critical']]]
    for r_idx, row_data in enumerate(chart_data_rows, 1):
        for c_idx, cell_data in enumerate(row_data, 20): ws_dash.cell(row=r_idx, column=c_idx, value=cell_data)
    labels = Reference(ws_dash, min_col=20, min_row=2, max_row=4); data = Reference(ws_dash, min_col=21, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True); pie.set_categories(labels)
    series = pie.series[0]; pts = [DataPoint(idx=i) for i in range(len(chart_data_rows) - 1)]
    slice_colors = ["C6EFCE", "FFEB9C", "FFC7CE"]
    for i, pt in enumerate(pts): pt.graphicalProperties.solidFill = slice_colors[i]
    series.dps = pts
    ws_dash.add_chart(pie, "B10")
    
    bar_chart = BarChart(); bar_chart.type = "col"; bar_chart.style = 10; bar_chart.title = "Основные точки отказа"; bar_chart.height = 10; bar_chart.width = 17
    bar_chart.y_axis.title = 'Количество ПК'
    problem_data = stats['problem_counts'].most_common(5)
    if problem_data:
        problem_chart_rows = [['Проблема', 'Кол-во']] + problem_data
        for r_idx, row_data in enumerate(problem_chart_rows, 1):
            for c_idx, cell_data in enumerate(row_data, 23): ws_dash.cell(row=r_idx, column=c_idx, value=cell_data)
        data = Reference(ws_dash, min_col=24, min_row=1, max_row=len(problem_chart_rows)); cats = Reference(ws_dash, min_col=23, min_row=2, max_row=len(problem_chart_rows))
        bar_chart.add_data(data, titles_from_data=True); bar_chart.set_categories(cats)
        ws_dash.add_chart(bar_chart, "H10")

    # --- ИСПРАВЛЕНО: Секция Топ-5 ОПУЩЕНА НИЖЕ ---
    ws_dash['B30'] = 'Приоритетные компьютеры'; ws_dash['B30'].font = section_title_font; ws_dash.merge_cells('B30:K30')
    ws_dash['B31'] = 'Имя ПК'; ws_dash['B31'].font = header_font; ws_dash['B31'].fill = header_fill
    ws_dash.merge_cells('C31:K31'); ws_dash['C31'] = 'Основные проблемы'; ws_dash['C31'].font = header_font; ws_dash['C31'].fill = header_fill
    cat1_fill = PatternFill(start_color="FFC7CE", fill_type="solid"); cat2_fill = PatternFill(start_color="FFEB9C", fill_type="solid"); cat3_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    for i, pc_data in enumerate(stats['top_5_critical'], 32):
        ws_dash.cell(row=i, column=2, value=pc_data.get('Название ПК', 'N/A'))
        ws_dash.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)
        ws_dash.cell(row=i, column=3, value=pc_data.get('problems', 'Нет данных'))
        cat_fill = {1: cat1_fill, 2: cat2_fill, 3: cat3_fill}.get(pc_data.get('category', 3))
        for cell_col in range(2, 12):
            cell = ws_dash.cell(row=i, column=cell_col)
            if cat_fill: cell.fill = cat_fill
            cell.border = thin_border
    
    for col_letter, width in [('A', 2), ('B', 22), ('C', 15), ('D', 15), ('E', 15), ('F', 15), ('G', 15), ('H', 15), ('I', 15), ('J', 15), ('K', 15)]:
        ws_dash.column_dimensions[col_letter].width = width

    # ОСТАЛЬНЫЕ ЛИСТЫ
    ws_main = wb.create_sheet("Все данные"); data_font = Font(name='Calibri', size=11); smart_bad_font = Font(bold=True, color="9C0006"); data_alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
    final_headers = [h for h in HEADERS_MAIN if h != '_RAW_DATA']; [final_headers.append(h) for h in HEADERS_NETWORK if h not in final_headers]
    ws_main.append(final_headers)
    for cell in ws_main[1]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
    for data_row in data_list:
        ws_main.append([data_row.get(h, '') for h in final_headers]); row_idx = ws_main.max_row
        category = data_row.get('category', 3); row_fill, row_font = {1: (cat1_fill, Font(color="9C0006")), 2: (cat2_fill, Font(color="9C6500")), 3: (cat3_fill, data_font)}.get(category, (None, data_font))
        for cell in ws_main[row_idx]: cell.font = row_font; cell.alignment = data_alignment; cell.border = thin_border
        if row_fill:
            for cell in ws_main[row_idx]: cell.fill = row_fill
        try:
            if data_row.get('internal_smart_status') == 'BAD': ws_main.cell(row=row_idx, column=final_headers.index('SMART Статус') + 1).font = smart_bad_font
        except (ValueError, IndexError): pass
    for i, header_text in enumerate(final_headers, 1):
        column_letter = get_column_letter(i); max_length = max([len(str(c.value)) for c in ws_main[column_letter] if c.value] or [0])
        ws_main.column_dimensions[column_letter].width = min(max(max_length, len(header_text)) + 2, 60)
    ws_main.freeze_panes = 'A2'; ws_analysis = wb.create_sheet("Рекомендации"); cat_font = Font(bold=True, name='Calibri', size=12)
    for cat_num, cat_name, cat_fill in [(1, 'Полная замена', cat1_fill), (2, 'Частичный апгрейд', cat2_fill)]:
        ws_analysis.append([f'Категория {cat_num}: {cat_name}']); merged_cell_header = ws_analysis.cell(row=ws_analysis.max_row, column=1)
        merged_cell_header.font = cat_font; merged_cell_header.fill = cat_fill; ws_analysis.merge_cells(f'A{ws_analysis.max_row}:{get_column_letter(len(HEADERS_ANALYSIS))}{ws_analysis.max_row}')
        ws_analysis.append(HEADERS_ANALYSIS)
        for cell in ws_analysis[ws_analysis.max_row]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        for data in data_list:
            if data.get('category') == cat_num:
                rec_list = set()
                if data.get('internal_smart_status') == "BAD": rec_list.add("ЗАМЕНА ДИСКА!")
                elif cat_num == 1: rec_list.add("Полная замена")
                else:
                    for p in str(data.get('problems', '')).lower().split('; '):
                        if 'ос' in p or 'windows 7' in p: rec_list.add("Обновить ОС")
                        elif 'ssd' in p: rec_list.add("Установить SSD")
                        elif 'озу' in p: rec_list.add("Добавить ОЗУ")
                        elif 'видеодрайвер' in p: rec_list.add("Установить видеодрайвер")
                        elif 'bios' in p: rec_list.add("Обновить BIOS (опционально)")
                ws_analysis.append([data.get(h, '') for h in ['Имя файла', 'Название ПК', 'problems']] + [", ".join(sorted(list(rec_list))) or "Частичный апгрейд"])
        ws_analysis.append([])
    for col_idx in range(1, ws_analysis.max_column + 1):
        col_letter = get_column_letter(col_idx); ws_analysis.column_dimensions[col_letter].width = max((len(str(c.value)) for c in ws_analysis[col_letter] if c.value and not isinstance(c, MergedCell)), default=20) + 2
    try: wb.save(filename); log_emitter(f"Файл Excel '{filename}' с дашбордом успешно сохранен.", "info")
    except IOError as e: log_emitter(f"Ошибка: Не удалось записать в файл {filename}. Возможно, он открыт. Ошибка: {e}", "error")